"""
Endpoints para facturas
"""
from fastapi import APIRouter, Depends, Query, HTTPException, Request
from fastapi.responses import Response
from starlette.requests import Request
from typing import Optional, Union
from datetime import date
from pydantic import BaseModel, Field
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.api.dependencies import get_factura_repository
from src.logging_conf import get_logger

logger = get_logger(__name__)
from src.api.schemas.facturas import (
    FacturaSummaryResponse,
    FacturaByDayResponse,
    FacturaByDayItem,
    FacturaRecentResponse,
    FacturaRecentItem,
    CategoryBreakdownResponse,
    CategoryBreakdownItem,
    FailedInvoicesResponse,
    FailedInvoiceItem,
    FacturaListResponse,
    FacturaListItem,
    ManualFacturaCreate,
)
from src.db.repositories import FacturaRepository

router = APIRouter(prefix="/facturas", tags=["facturas"])


@router.get("/summary", response_model=FacturaSummaryResponse)
async def get_summary(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener resumen de facturas del mes seleccionado
    """
    try:
        summary = repo.get_summary_by_month(month, year)
        return FacturaSummaryResponse(**summary)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener resumen: {str(e)}")


@router.get("/by_day", response_model=FacturaByDayResponse)
async def get_by_day(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener facturas agrupadas por día del mes
    """
    try:
        by_day = repo.get_facturas_by_day(month, year)
        items = [FacturaByDayItem(**item) for item in by_day]
        return FacturaByDayResponse(data=items)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener datos por día: {str(e)}")


@router.get("/recent", response_model=FacturaRecentResponse)
async def get_recent(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    limit: int = Query(5, ge=1, le=100, description="Límite de facturas"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener facturas recientes del mes
    """
    try:
        recent = repo.get_recent_facturas(month, year, limit)
        items = []
        for item in recent:
            # Convertir fecha_emision a date si es datetime
            fecha_emision = item.get('fecha_emision')
            if fecha_emision and hasattr(fecha_emision, 'date'):
                fecha_emision = fecha_emision.date()
            elif fecha_emision and isinstance(fecha_emision, str):
                from datetime import datetime
                fecha_emision = datetime.fromisoformat(fecha_emision).date()
            
            items.append(FacturaRecentItem(
                id=item['id'],
                numero_factura=item.get('numero_factura'),
                proveedor_nombre=item.get('proveedor_nombre'),
                fecha_emision=fecha_emision,
                importe_base=item.get('importe_base'),
                importe_iva=item.get('importe_iva'),
                importe_total=item.get('importe_total')
            ))
        return FacturaRecentResponse(data=items)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener facturas recientes: {str(e)}")


@router.get("/list", response_model=FacturaListResponse)
async def get_all_facturas(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener todas las facturas del mes
    """
    try:
        facturas = repo.get_all_facturas_by_month(month, year)
        items = []
        for item in facturas:
            items.append(FacturaListItem(**item))
        
        return FacturaListResponse(data=items)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener facturas: {str(e)}")


@router.get("/categories", response_model=CategoryBreakdownResponse)
async def get_categories(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener desglose por categorías (proveedores)
    """
    try:
        categories = repo.get_categories_breakdown(month, year)
        items = [CategoryBreakdownItem(**item) for item in categories]
        return CategoryBreakdownResponse(data=items)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener desglose por categorías: {str(e)}")


@router.get("/failed", response_model=FailedInvoicesResponse)
async def get_failed_invoices(
    request: Request,
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Obtener lista de facturas fallidas.
    Si se proporcionan month y year como query params, filtra por ese mes. Si no, devuelve TODAS las facturas fallidas.
    Combina facturas en BD (estado 'error' o 'revisar') con archivos en cuarentena.
    """
    try:
        import json
        import os
        import re
        from pathlib import Path
        from datetime import datetime, date
        from calendar import monthrange
        
        # Leer query params manualmente
        query_params = request.query_params
        month_str = query_params.get('month')
        year_str = query_params.get('year')
        
        month = None
        year = None
        
        if month_str:
            try:
                month = int(month_str)
                if month < 1 or month > 12:
                    raise HTTPException(status_code=422, detail="month debe estar entre 1 y 12")
            except (ValueError, TypeError):
                raise HTTPException(status_code=422, detail="month debe ser un número entero entre 1 y 12")
        
        if year_str:
            try:
                year = int(year_str)
                if year < 2000 or year > 2100:
                    raise HTTPException(status_code=422, detail="year debe estar entre 2000 y 2100")
            except (ValueError, TypeError):
                raise HTTPException(status_code=422, detail="year debe ser un número entero entre 2000 y 2100")
        
        # Si se proporcionan month y year, calcular rango de fechas
        # Si no, usar None para indicar que no hay filtro
        start_date = None
        end_date = None
        if month is not None and year is not None:
            start_date = date(year, month, 1)
            _, last_day = monthrange(year, month)
            end_date = date(year, month, last_day)
        
        failed_invoices = []
        processed_names_bd = set()  # Solo para evitar duplicados en BD
        processed_names_quarantine = set()  # Para evitar duplicados entre archivos de cuarentena
        
        logger.info(f"🔍 Iniciando búsqueda de facturas fallidas. Filtro: month={month}, year={year}")
        
        # 1. CONSULTAR FACTURAS EN BD CON ESTADO "error" o "revisar"
        logger.info("📊 Paso 1: Consultando facturas en BD con estado 'error' o 'revisar'")
        with repo.db.get_session() as session:
            from src.db.models import Factura
            
            # Query base: facturas con estado error o revisar
            query = session.query(Factura).filter(
                Factura.estado.in_(['error', 'revisar'])
            )
            
            # Si hay filtro de fecha, aplicarlo
            if start_date is not None and end_date is not None:
                from sqlalchemy import func
                fecha_filtro = func.coalesce(Factura.fecha_emision, Factura.fecha_recepcion)
                query = query.filter(
                    fecha_filtro >= start_date,
                    fecha_filtro <= end_date
                )
                logger.info(f"📅 Aplicando filtro de fecha: {start_date} a {end_date}")
            
            facturas_bd = query.all()
            logger.info(f"✅ Encontradas {len(facturas_bd)} facturas en BD con estado 'error' o 'revisar'")
            
            for factura_obj in facturas_bd:
                nombre = factura_obj.drive_file_name
                if nombre and nombre not in processed_names_bd:
                    # Extraer razón del error_msg si existe
                    razon = factura_obj.error_msg if hasattr(factura_obj, 'error_msg') and factura_obj.error_msg else None
                    if not razon:
                        # Si no hay error_msg, usar el estado como razón
                        razon = f"Estado: {factura_obj.estado}"
                    
                    failed_invoices.append({
                        'nombre': nombre,
                        'fecha_emision': factura_obj.fecha_emision.isoformat() if factura_obj.fecha_emision else None,
                        'estado': factura_obj.estado,
                        'source': 'bd',
                        'razon': razon
                    })
                    processed_names_bd.add(nombre)
                    logger.debug(f"  ➕ Agregada factura de BD: {nombre} (estado: {factura_obj.estado})")
        
        logger.info(f"📊 Total facturas de BD agregadas: {len(failed_invoices)}")
        
        # 2. CONSULTAR ARCHIVOS EN CUARENTENA
        logger.info("📁 Paso 2: Consultando archivos en cuarentena")
        # Usar ruta relativa al directorio de trabajo del contenedor (/app)
        quarantine_path_env = os.getenv('QUARANTINE_PATH', 'data/quarantine')
        
        # Siempre usar ruta relativa desde /app (directorio de trabajo del contenedor)
        # Ignorar si viene como ruta absoluta del host
        if os.path.isabs(quarantine_path_env):
            # Si viene como ruta absoluta, extraer solo el nombre relativo
            # Ej: /home/alex/proyectos/invoice-extractor/data/quarantine -> data/quarantine
            parts = Path(quarantine_path_env).parts
            if 'data' in parts and 'quarantine' in parts:
                # Encontrar índice de 'data' y tomar desde ahí
                data_idx = parts.index('data')
                quarantine_path_env = str(Path(*parts[data_idx:]))
                logger.info(f"📂 Ruta absoluta detectada, convirtiendo a relativa: {quarantine_path_env}")
        
        # Resolver desde /app (directorio de trabajo del contenedor)
        quarantine_path = Path('/app') / quarantine_path_env
        logger.info(f"📂 Ruta de cuarentena (env): {os.getenv('QUARANTINE_PATH', 'data/quarantine')}")
        logger.info(f"📂 Ruta de cuarentena (usada): {quarantine_path_env}")
        logger.info(f"📂 Ruta de cuarentena (resuelta): {quarantine_path}")
        logger.info(f"📂 Existe: {quarantine_path.exists()}")
        
        # Obtener lista de facturas ya procesadas en BD para filtrar cuarentena
        processed_in_bd = set()
        with repo.db.get_session() as session:
            from src.db.models import Factura
            # Obtener todas las facturas procesadas (estado diferente de 'error' y 'revisar')
            facturas_procesadas = session.query(Factura.drive_file_name).filter(
                Factura.drive_file_name.isnot(None),
                ~Factura.estado.in_(['error', 'revisar'])
            ).all()
            processed_in_bd = {f[0] for f in facturas_procesadas if f[0]}
            logger.info(f"📊 Facturas ya procesadas en BD (excluidas de cuarentena): {len(processed_in_bd)}")
        
        quarantine_stats = {
            'total_files_found': 0,
            'processed': 0,
            'skipped_bd_duplicate': 0,
            'skipped_quarantine_duplicate': 0,
            'skipped_unknown': 0,
            'errors': 0,
            'error_details': []
        }
        
        if quarantine_path.exists():
            # Buscar en todas las subcarpetas de cuarentena
            meta_files = list(quarantine_path.rglob("*.meta.json"))
            quarantine_stats['total_files_found'] = len(meta_files)
            logger.info(f"🔍 Encontrados {len(meta_files)} archivos .meta.json en cuarentena")
            
            for meta_file in meta_files:
                try:
                    with open(meta_file, 'r', encoding='utf-8') as f:
                        meta_data = json.load(f)
                    
                    # Extraer información del archivo - múltiples fuentes posibles
                    file_info = meta_data.get('file_info', {})
                    nombre = (
                        meta_data.get('drive_file_name') or 
                        file_info.get('name') or 
                        meta_file.stem.replace('.meta', '').split('_', 2)[-1] if '_' in meta_file.stem else meta_file.stem.replace('.meta', '')
                    )
                    
                    # Omitir si está en BD o si ya fue procesado de cuarentena (evitar duplicados)
                    if nombre == 'unknown':
                        quarantine_stats['skipped_unknown'] += 1
                        logger.debug(f"  ⏭️  Omitido (nombre unknown): {meta_file.name}")
                        continue
                    
                    if nombre in processed_names_bd:
                        quarantine_stats['skipped_bd_duplicate'] += 1
                        logger.debug(f"  ⏭️  Omitido (ya en BD con error/revisar): {nombre}")
                        continue
                    
                    # Excluir facturas que ya fueron procesadas correctamente en BD
                    if nombre in processed_in_bd:
                        quarantine_stats['skipped_bd_duplicate'] += 1
                        logger.debug(f"  ⏭️  Omitido (ya procesada en BD): {nombre}")
                        continue
                    
                    # Evitar duplicados entre archivos de cuarentena (solo incluir cada factura única una vez)
                    if nombre in processed_names_quarantine:
                        quarantine_stats['skipped_quarantine_duplicate'] += 1
                        logger.debug(f"  ⏭️  Omitido (duplicado en cuarentena): {nombre}")
                        continue
                    
                    # Intentar obtener fecha_emision del metadata (si fue extraída antes de fallar)
                    file_date = None
                    fecha_emision_str = meta_data.get('fecha_emision') or meta_data.get('factura_data', {}).get('fecha_emision')
                    
                    if fecha_emision_str:
                        try:
                            if isinstance(fecha_emision_str, str):
                                file_date = datetime.fromisoformat(fecha_emision_str.replace('Z', '+00:00')).date()
                            else:
                                # Si es un objeto date/datetime
                                if hasattr(fecha_emision_str, 'date'):
                                    file_date = fecha_emision_str.date()
                                elif hasattr(fecha_emision_str, 'isoformat'):
                                    file_date = datetime.fromisoformat(fecha_emision_str.isoformat()).date()
                        except (ValueError, AttributeError, TypeError):
                            pass
                    
                    # Si no hay fecha_emision en metadata, intentar parsear del nombre del archivo
                    # Esto es crítico porque muchos archivos en cuarentena no tienen fecha_emision
                    if file_date is None:
                        default_year = year if year is not None else datetime.utcnow().year
                        file_date = _parse_date_from_filename(nombre, default_year)
                    
                    # Si aún no tenemos fecha, usar fecha de modificación de Drive como siguiente opción
                    if file_date is None:
                        modified_time = file_info.get('modifiedTime') or meta_data.get('file_info', {}).get('modifiedTime')
                        if modified_time:
                            try:
                                if isinstance(modified_time, str):
                                    file_date = datetime.fromisoformat(modified_time.replace('Z', '+00:00')).date()
                            except (ValueError, AttributeError):
                                pass
                    
                    # Si aún no tenemos fecha, usar fecha de cuarentena como último recurso
                    if file_date is None:
                        quarantined_at_str = meta_data.get('quarantined_at')
                        if quarantined_at_str:
                            try:
                                file_date = datetime.fromisoformat(quarantined_at_str.replace('Z', '+00:00')).date()
                            except (ValueError, AttributeError):
                                pass
                    
                    # Si hay filtro de fecha, verificar que la fecha esté en el rango
                    # Si no hay filtro, incluir todas las facturas
                    should_include = False
                    
                    if start_date is not None and end_date is not None:
                        # Hay filtro de mes: solo incluir si la fecha está en el rango
                        if file_date is not None:
                            if start_date <= file_date <= end_date:
                                should_include = True
                        # Si NO tenemos fecha pero el nombre sugiere el mes correcto, incluirlo de todas formas
                        elif nombre and nombre != 'unknown':
                            # Verificar si el nombre contiene el mes en texto
                            nombre_lower = nombre.lower()
                            meses_nombres = {
                                1: ['enero', 'jan'], 2: ['febrero', 'feb'], 3: ['marzo', 'mar'],
                                4: ['abril', 'apr'], 5: ['mayo', 'may'], 6: ['junio', 'jun'],
                                7: ['julio', 'jul'], 8: ['agosto', 'agost', 'aug'], 9: ['septiembre', 'sep', 'sept'],
                                10: ['octubre', 'oct'], 11: ['noviembre', 'nov'], 12: ['diciembre', 'dec']
                            }
                            
                            # Verificar si el nombre contiene el mes actual
                            mes_encontrado = False
                            for mes_nombre in meses_nombres.get(month, []):
                                if mes_nombre in nombre_lower:
                                    # Verificar si también tiene el año
                                    if str(year) in nombre or str(year)[-2:] in nombre:
                                        mes_encontrado = True
                                        break
                            
                            if mes_encontrado:
                                should_include = True
                    else:
                        # No hay filtro: incluir todas las facturas en cuarentena
                        should_include = True
                    
                    if should_include:
                        # Incluir TODAS las entradas de cuarentena (sin deduplicar)
                        # Agregar información adicional para diferenciar duplicados
                        # Extraer razón: priorizar 'reason', luego 'error', luego 'decision'
                        razon = (
                            meta_data.get('reason') or 
                            meta_data.get('error') or 
                            (f"Decisión: {meta_data.get('decision')}" if meta_data.get('decision') else None)
                        )
                        # Limpiar mensajes de error muy largos (ej: stack traces de BD)
                        if razon and len(razon) > 200:
                            razon = razon[:200] + "..."
                        
                        quarantine_entry = {
                            'nombre': nombre,
                            'fecha_emision': file_date.isoformat() if file_date else None,
                            'estado': 'quarantine',
                            'source': 'quarantine',
                            'meta_file': str(meta_file.relative_to(quarantine_path)) if quarantine_path in meta_file.parents else meta_file.name,
                            'quarantined_at': meta_data.get('quarantined_at'),
                            'decision': meta_data.get('decision'),
                            'razon': razon
                        }
                        failed_invoices.append(quarantine_entry)
                        processed_names_quarantine.add(nombre)  # Marcar como procesado para evitar duplicados
                        quarantine_stats['processed'] += 1
                        logger.debug(f"  ➕ Agregada factura de cuarentena: {nombre} (archivo: {meta_file.name})")
                
                except json.JSONDecodeError as e:
                    quarantine_stats['errors'] += 1
                    error_msg = f"Error JSON en {meta_file.name}: {str(e)}"
                    quarantine_stats['error_details'].append(error_msg)
                    logger.warning(f"  ⚠️  {error_msg}")
                    continue
                except ValueError as e:
                    quarantine_stats['errors'] += 1
                    error_msg = f"Error de valor en {meta_file.name}: {str(e)}"
                    quarantine_stats['error_details'].append(error_msg)
                    logger.warning(f"  ⚠️  {error_msg}")
                    continue
                except KeyError as e:
                    quarantine_stats['errors'] += 1
                    error_msg = f"Error de clave faltante en {meta_file.name}: {str(e)}"
                    quarantine_stats['error_details'].append(error_msg)
                    logger.warning(f"  ⚠️  {error_msg}")
                    continue
                except Exception as e:
                    quarantine_stats['errors'] += 1
                    error_msg = f"Error inesperado procesando {meta_file.name}: {type(e).__name__}: {str(e)}"
                    quarantine_stats['error_details'].append(error_msg)
                    logger.error(f"  ❌ {error_msg}", exc_info=True)
                    continue
        else:
            logger.warning(f"⚠️  La ruta de cuarentena no existe: {quarantine_path}")
        
        logger.info(f"📊 Estadísticas de cuarentena:")
        logger.info(f"  - Archivos encontrados: {quarantine_stats['total_files_found']}")
        logger.info(f"  - Procesados (únicos): {quarantine_stats['processed']}")
        logger.info(f"  - Omitidos (duplicados de BD): {quarantine_stats['skipped_bd_duplicate']}")
        logger.info(f"  - Omitidos (duplicados en cuarentena): {quarantine_stats['skipped_quarantine_duplicate']}")
        logger.info(f"  - Omitidos (nombre unknown): {quarantine_stats['skipped_unknown']}")
        logger.info(f"  - Errores: {quarantine_stats['errors']}")
        if quarantine_stats['error_details']:
            logger.warning(f"  - Detalles de errores: {quarantine_stats['error_details'][:5]}")  # Primeros 5 errores
        
        # Ordenar por fecha_emision (más recientes primero)
        # Si no hay fecha, ponerlas al final
        failed_invoices.sort(
            key=lambda x: x.get('fecha_emision', '') or '0000-00-00',
            reverse=True
        )
        
        logger.info(f"📊 Total facturas fallidas encontradas: {len(failed_invoices)}")
        logger.info(f"  - De BD: {sum(1 for item in failed_invoices if item.get('source') == 'bd')}")
        logger.info(f"  - De cuarentena: {sum(1 for item in failed_invoices if item.get('source') == 'quarantine')}")
        
        # Limitar a 500 resultados para evitar sobrecarga (aumentado de 200 para incluir todas las entradas)
        if len(failed_invoices) > 500:
            logger.warning(f"⚠️  Limitando resultados a 500 (había {len(failed_invoices)})")
            failed_invoices = failed_invoices[:500]
        
        # Construir items con nombre y razón
        items = [FailedInvoiceItem(
            nombre=item['nombre'],
            razon=item.get('razon')
        ) for item in failed_invoices]
        
        logger.info(f"✅ Devolviendo {len(items)} facturas fallidas")
        return FailedInvoicesResponse(data=items)
    
    except HTTPException:
        raise  # Re-raise HTTPException sin modificar
    except Exception as e:
        logger.error(f"❌ Error inesperado en get_failed_invoices: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al obtener facturas fallidas: {str(e)}")


def _parse_date_from_filename(filename: str, default_year: int = None) -> Optional[date]:
    """
    Intentar parsear fecha desde el nombre del archivo.
    Ejemplos: "Factura REVO 1 Enero 2024" -> 2024-01-01
              "Fact EVOLBE jul 25" -> 2025-07-25 (si default_year=2025)
              "Factura REVO 2 Enero 2024.pdf" -> 2024-01-01
    """
    try:
        from datetime import datetime
        import dateparser
        
        # Mapeo de meses en español (incluyendo variantes)
        meses_es = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
            'julio': 7, 'agosto': 8, 'agost': 8,  # Variante común
            'septiembre': 9, 'sep': 9, 'sept': 9,
            'octubre': 10, 'oct': 10,
            'noviembre': 11, 'nov': 11,
            'diciembre': 12, 'dec': 12,
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        
        # Normalizar nombre: remover extensión y convertir a minúsculas para búsqueda
        filename_clean = filename.rsplit('.', 1)[0].lower()
        
        # Patrones mejorados para nombres de archivos
        patterns = [
            # "Enero 2024", "enero 2024" - patrón más flexible
            r'(enero|febrero|marzo|abril|mayo|junio|julio|agosto|agost|septiembre|sep|sept|octubre|oct|noviembre|nov|diciembre|dec)\s+(\d{4})',
            # "jul 25", "julio 25", "agost 25" - con año de 2 dígitos
            r'(enero|febrero|marzo|abril|mayo|junio|julio|agosto|agost|septiembre|sep|sept|octubre|oct|noviembre|nov|diciembre|dec)\s+(\d{2})',
            # "2024-01", "2024/01"
            r'(\d{4})[-/](\d{1,2})',
            # Patrón específico: "Factura X Enero 2024" o "Fact X jul 25"
            r'(?:factura|fact)\s+\w+\s+\d*\s*(enero|febrero|marzo|abril|mayo|junio|julio|agosto|agost|septiembre|sep|sept|octubre|oct|noviembre|nov|diciembre|dec)\s+(\d{2,4})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename_clean, re.IGNORECASE)
            if match:
                grupos = match.groups()
                
                # Patrón: mes + año (4 dígitos)
                if len(grupos) == 2:
                    grupo1 = grupos[0].lower()
                    grupo2 = grupos[1]
                    
                    # Si grupo1 es un mes y grupo2 es un año (4 dígitos)
                    if grupo1 in meses_es and len(grupo2) == 4:
                        year = int(grupo2)
                        month = meses_es[grupo1]
                        return date(year, month, 1)
                    
                    # Si grupo1 es un mes y grupo2 es un año (2 dígitos)
                    if grupo1 in meses_es and len(grupo2) == 2:
                        year = 2000 + int(grupo2)  # Asumir 20XX
                        month = meses_es[grupo1]
                        return date(year, month, 1)
                    
                    # Si grupo1 es año y grupo2 es mes
                    if len(grupo1) == 4 and grupo2.isdigit():
                        year = int(grupo1)
                        month = int(grupo2)
                        if 1 <= month <= 12:
                            return date(year, month, 1)
        
        # Intentar con dateparser como último recurso
        if dateparser:
            try:
                # Usar default_year si está disponible
                settings = {}
                if default_year:
                    settings['PREFER_DATES_FROM'] = 'future'
                    settings['RELATIVE_BASE'] = datetime(default_year, 1, 1)
                
                parsed = dateparser.parse(filename, languages=['es', 'en'], settings=settings)
                if parsed:
                    return parsed.date()
            except:
                pass
        
        return None
    except Exception:
        return None



@router.get("/export/excel/pendientes")
async def export_facturas_pendientes_to_excel(
    request: Request,
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Exportar todas las facturas pendientes (con errores o en cuarentena) a un archivo Excel bien formateado.
    Combina facturas en BD (estado 'error') con archivos en cuarentena (igual que la tabla de Pendientes).
    """
    try:
        import json
        import os
        import re
        from pathlib import Path
        from datetime import datetime, date
        from calendar import monthrange
        
        # Leer query params opcionales (no obligatorios)
        query_params = request.query_params
        month_str = query_params.get('month')
        year_str = query_params.get('year')
        
        month = None
        year = None
        
        if month_str:
            try:
                month = int(month_str)
            except (ValueError, TypeError):
                pass
        
        if year_str:
            try:
                year = int(year_str)
            except (ValueError, TypeError):
                pass
        
        # Calcular rango de fechas si se proporcionaron
        start_date = None
        end_date = None
        if month is not None and year is not None:
            start_date = date(year, month, 1)
            _, last_day = monthrange(year, month)
            end_date = date(year, month, last_day)
        
        facturas_data = []
        processed_names = set()  # Para evitar duplicados (inicializar al principio)
        
        logger.info(f"🔍 Iniciando export de facturas pendientes")
        
        # 1. Obtener facturas de BD con estado 'error' o 'error_permanente'
        with repo.db.get_session() as session:
            from src.db.models import Factura
            from sqlalchemy import func, or_
            
            query = session.query(Factura).filter(
                or_(
                    Factura.estado == 'error',
                    Factura.estado == 'error_permanente'
                )
            )
            
            # Filtrar por fecha si se proporcionó
            if start_date and end_date:
                fecha_filtro = func.coalesce(Factura.fecha_emision, Factura.fecha_recepcion)
                query = query.filter(
                    fecha_filtro >= start_date,
                    fecha_filtro <= end_date
                )
            
            facturas_bd = query.all()
            
            for f in facturas_bd:
                # Filtrar duplicados también en BD
                nombre = f.drive_file_name or f'factura_{f.id}'
                if nombre in processed_names:
                    continue
                processed_names.add(nombre)
                
                facturas_data.append({
                    'id': f.id,
                    'nombre_archivo': f.drive_file_name or f'Factura {f.id}',
                    'proveedor_nombre': f.proveedor_text or 'Sin proveedor',
                    'categoria': f.drive_folder_name or 'Sin categoría',
                    'fecha_emision': f.fecha_emision.isoformat() if f.fecha_emision else 'Sin fecha',
                    'estado': 'BD - Error',
                    'motivo_error': f.motivo_rechazo if hasattr(f, 'motivo_rechazo') else 'Error en procesamiento',
                    'impuestos_total': float(f.impuestos_total) if f.impuestos_total else 0.0,
                    'importe_total': float(f.importe_total) if f.importe_total else 0.0
                })
        
        logger.info(f"📊 Facturas desde BD con error: {len(facturas_bd)} (agregadas: {len(facturas_data)})")
        
        # 2. Obtener lista de facturas ya procesadas CORRECTAMENTE en BD (para excluir de cuarentena)
        processed_in_bd = set()
        
        with repo.db.get_session() as session:
            from src.db.models import Factura
            # Obtener SOLO las facturas con estado 'procesado' (no las con error)
            facturas_procesadas = session.query(Factura.drive_file_name).filter(
                Factura.drive_file_name.isnot(None),
                Factura.estado == 'procesado'
            ).all()
            processed_in_bd = {f[0] for f in facturas_procesadas if f[0]}
        
        logger.info(f"📋 Facturas procesadas correctamente en BD: {len(processed_in_bd)}")
        
        # 3. Obtener facturas de cuarentena (archivos que fallaron)
        quarantine_path = Path(os.getenv('QUARANTINE_PATH', 'data/quarantine'))
        if quarantine_path.exists():
            # Buscar archivos .meta.json en todas las subcarpetas de cuarentena
            meta_files = list(quarantine_path.rglob("*.meta.json"))
            logger.info(f"📁 Archivos .meta.json encontrados en cuarentena: {len(meta_files)}")
            
            excluidos_procesados = 0
            excluidos_duplicados = 0
            
            for meta_file in meta_files:
                try:
                    # Leer metadata
                    with open(meta_file, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                    
                    # Extraer información
                    file_info = metadata.get('file_info', {})
                    nombre = (
                        metadata.get('drive_file_name') or 
                        file_info.get('name') or 
                        meta_file.stem.replace('.meta', '')
                    )
                    
                    # FILTRO 1: Omitir si ya fue procesada correctamente en BD
                    if nombre in processed_in_bd:
                        excluidos_procesados += 1
                        continue
                    
                    # FILTRO 2: Omitir si ya la agregamos (evitar duplicados)
                    if nombre in processed_names:
                        excluidos_duplicados += 1
                        continue
                    
                    processed_names.add(nombre)
                    
                    # Extraer fecha de file_info.modifiedTime
                    fecha_str = 'Sin fecha'
                    modified_time = file_info.get('modifiedTime')
                    if modified_time:
                        try:
                            # Parsear fecha ISO8601 (ej: "2025-05-02T19:11:42.463Z")
                            fecha_obj = datetime.fromisoformat(modified_time.replace('Z', '+00:00')).date()
                            fecha_str = fecha_obj.isoformat()
                            
                            # Filtrar por mes/año si se proporcionó
                            if start_date and end_date:
                                if not (start_date <= fecha_obj <= end_date):
                                    continue
                        except:
                            pass
                    
                    # Extraer motivo del rechazo
                    motivo = metadata.get('reason', 'Error al procesar archivo')
                    
                    # Extraer datos de factura_data si existe
                    factura_data = metadata.get('factura_data', {})
                    
                    # Proveedor: intentar de factura_data, sino derivar del nombre del archivo
                    proveedor = factura_data.get('proveedor_text')
                    if not proveedor:
                        # Intentar derivar del nombre del archivo
                        # Ejemplo: "Fact MAKRO 1 may 25.pdf" → "MAKRO"
                        nombre_upper = nombre.upper()
                        if 'MAKRO' in nombre_upper:
                            proveedor = 'Makro'
                        elif 'NEGRINI' in nombre_upper:
                            proveedor = 'Negrini'
                        elif 'CONWAY' in nombre_upper:
                            proveedor = 'Conway'
                        elif 'REVO' in nombre_upper:
                            proveedor = 'Revo'
                        elif 'CAFENTO' in nombre_upper:
                            proveedor = 'Cafento'
                        elif 'QUIRON' in nombre_upper or 'QUIRONPREVENCION' in nombre_upper:
                            proveedor = 'Quirón Prevención'
                        else:
                            proveedor = 'Sin proveedor'
                    
                    # Categoría: intentar derivar del tipo de documento
                    categoria = metadata.get('drive_folder_name')
                    if not categoria:
                        categoria = file_info.get('folder_name')
                    if not categoria or categoria in ['review', 'duplicates', 'otros']:
                        # Intentar derivar del nombre
                        nombre_lower = nombre.lower()
                        if any(x in nombre_lower for x in ['luz', 'energia', 'electric']):
                            categoria = 'Servicios/Luz'
                        elif any(x in nombre_lower for x in ['agua', 'water']):
                            categoria = 'Servicios/Agua'
                        elif any(x in nombre_lower for x in ['telefon', 'internet', 'movil']):
                            categoria = 'Servicios/Telecomunicaciones'
                        elif any(x in nombre_lower for x in ['alquiler', 'rent', 'arrendamiento']):
                            categoria = 'Alquiler'
                        elif any(x in nombre_lower for x in ['honorarios', 'abogado', 'notaria', 'gestor']):
                            categoria = 'Servicios Profesionales'
                        else:
                            categoria = 'Sin categoría'
                    
                    # Importes: parsear si son strings
                    importe_total = 0.0
                    impuestos_total = 0.0
                    
                    if factura_data.get('importe_total'):
                        try:
                            importe_total = float(str(factura_data['importe_total']).replace(',', '.'))
                        except (ValueError, TypeError):
                            importe_total = 0.0
                    
                    if factura_data.get('impuestos_total'):
                        try:
                            impuestos_total = float(str(factura_data['impuestos_total']).replace(',', '.'))
                        except (ValueError, TypeError):
                            impuestos_total = 0.0
                    
                    # Fecha: usar fecha_emision si existe, sino usar modifiedTime
                    if not fecha_str or fecha_str == 'Sin fecha':
                        fecha_emision_meta = factura_data.get('fecha_emision')
                        if fecha_emision_meta:
                            try:
                                fecha_obj_meta = datetime.fromisoformat(fecha_emision_meta).date()
                                fecha_str = fecha_obj_meta.isoformat()
                            except:
                                pass
                    
                    facturas_data.append({
                        'id': 'Cuarentena',
                        'nombre_archivo': nombre,
                        'proveedor_nombre': proveedor,
                        'categoria': categoria,
                        'fecha_emision': fecha_str,
                        'estado': 'Cuarentena',
                        'motivo_error': motivo,
                        'impuestos_total': impuestos_total,
                        'importe_total': importe_total
                    })
                except Exception as e:
                    logger.error(f"Error procesando archivo de cuarentena {meta_file}: {e}")
                    continue
            
            logger.info(f"🔍 Resultados de cuarentena: {len(meta_files)} archivos, {excluidos_procesados} excluidos (procesados), {excluidos_duplicados} excluidos (duplicados), {len(facturas_data) - len(facturas_bd)} agregados")
        
        logger.info(f"✅ Total facturas pendientes para Excel: {len(facturas_data)}")
        
        # Si no hay facturas, devolver Excel con mensaje
        if not facturas_data:
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_empty = pd.DataFrame({
                    'Mensaje': ['No se encontraron facturas pendientes']
                })
                sheet_name = "Facturas Pendientes"
                df_empty.to_excel(writer, index=False, sheet_name=sheet_name)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                cell = worksheet.cell(row=1, column=1)
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell = worksheet.cell(row=2, column=1)
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet.column_dimensions['A'].width = 60
            
            output.seek(0)
            
            file_name = "Facturas_Pendientes.xlsx"
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={file_name}"}
            )
        
        # LOG TEMPORAL: Ver cuántas facturas se están exportando
        logger.info(f"📊 EXPORTACIÓN PENDIENTES: Total facturas a exportar: {len(facturas_data)}")
        
        # Crear DataFrame y Excel
        df = pd.DataFrame(facturas_data)
        
        # Renombrar columnas
        df = df.rename(columns={
            'id': 'ID',
            'nombre_archivo': 'Nombre Archivo',
            'proveedor_nombre': 'Proveedor',
            'categoria': 'Categoría',
            'fecha_emision': 'Fecha Emisión',
            'estado': 'Estado',
            'motivo_error': 'Motivo',
            'impuestos_total': 'Impuestos',
            'importe_total': 'Total'
        })
        
        # Ordenar columnas
        df = df[['ID', 'Nombre Archivo', 'Proveedor', 'Categoría', 'Fecha Emisión', 'Estado', 'Motivo', 'Impuestos', 'Total']]
        
        # Crear archivo Excel
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = "Facturas Pendientes"
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos - aplicar formato SIN modificar valores (pandas ya los escribió)
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    
                    # Aplicar formato de moneda MANTENIENDO el valor existente
                    if df.columns[col_num - 1] in ['Impuestos', 'Total']:
                        cell.number_format = '#,##0.00 €'
                    elif df.columns[col_num - 1] == 'Fecha Emisión':
                        # Si la fecha es válida, formatearla
                        if isinstance(cell.value, str) and cell.value != 'Sin fecha':
                            try:
                                from datetime import datetime
                                fecha_obj = datetime.fromisoformat(cell.value)
                                cell.value = fecha_obj
                                cell.number_format = 'dd/mm/yyyy'
                            except:
                                pass  # Mantener como texto si no se puede parsear
            
            # Ajustar anchos
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(64 + i)].width = min(max_len, 50)
            
            # Fila de totales
            total_row_num = len(df) + 2
            total_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            total_font = Font(bold=True)
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=total_row_num, column=col_num)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                
                if df.columns[col_num - 1] == 'Proveedor':
                    cell.value = f'TOTAL: {len(df)} facturas pendientes'
                elif df.columns[col_num - 1] == 'Impuestos':
                    cell.value = df['Impuestos'].sum()
                    cell.number_format = '#,##0.00 €'
                elif df.columns[col_num - 1] == 'Total':
                    cell.value = df['Total'].sum()
                    cell.number_format = '#,##0.00 €'
        
        output.seek(0)
        
        file_name = "Facturas_Pendientes.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={file_name}",
                "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar facturas pendientes a Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar facturas pendientes a Excel: {str(e)}")


@router.get("/export/excel")
async def export_facturas_to_excel(
    month: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Año"),
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Exportar facturas del mes a Excel bien formateado
    """
    try:
        # Obtener todas las facturas del mes
        facturas = repo.get_all_facturas_by_month(month, year)
        
        if not facturas:
            raise HTTPException(status_code=404, detail=f"No hay facturas para {month}/{year}")
        
        # Convertir a DataFrame
        df = pd.DataFrame(facturas)
        
        # Renombrar columnas para mejor presentación
        column_mapping = {
            'id': 'ID',
            'proveedor_nombre': 'Proveedor',
            'categoria': 'Categoría',
            'fecha_emision': 'Fecha Emisión',
            'impuestos_total': 'Impuestos',
            'importe_total': 'Total'
        }
        
        # Seleccionar y renombrar columnas
        df_export = df[list(column_mapping.keys())].copy()
        df_export.rename(columns=column_mapping, inplace=True)
        
        # Formatear fecha
        if 'Fecha Emisión' in df_export.columns:
            df_export['Fecha Emisión'] = pd.to_datetime(df_export['Fecha Emisión'], errors='coerce')
            df_export['Fecha Emisión'] = df_export['Fecha Emisión'].dt.strftime('%d/%m/%Y')
        
        # Formatear números como moneda
        if 'Impuestos' in df_export.columns:
            df_export['Impuestos'] = df_export['Impuestos'].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "0.00")
        
        if 'Total' in df_export.columns:
            df_export['Total'] = df_export['Total'].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "0.00")
        
        # Crear Excel en memoria
        output = BytesIO()
        
        # Escribir DataFrame a Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Facturas', index=False)
            
            # Obtener el workbook y worksheet para formatear
            workbook = writer.book
            worksheet = writer.sheets['Facturas']
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            
            # Formatear encabezados
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border_style
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 8,   # ID
                'B': 30,  # Proveedor
                'C': 20,  # Categoría
                'D': 15,  # Fecha Emisión
                'E': 15,  # Impuestos
                'F': 15   # Total
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatear celdas de datos
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border_style
                    # Alinear números a la derecha
                    if cell.column_letter in ['E', 'F']:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = Alignment(vertical='center')
            
            # Agregar fila de totales
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=4, value="TOTAL:")
            worksheet.cell(row=total_row, column=4).font = Font(bold=True)
            worksheet.cell(row=total_row, column=4).alignment = right_alignment
            
            # Calcular totales (convertir strings de vuelta a números)
            total_impuestos = sum(float(str(x).replace(',', '')) if pd.notna(x) and str(x) != '' else 0 
                                 for x in df['impuestos_total'])
            total_importe = sum(float(str(x).replace(',', '')) if pd.notna(x) and str(x) != '' else 0 
                               for x in df['importe_total'])
            
            worksheet.cell(row=total_row, column=5, value=f"{total_impuestos:,.2f}")
            worksheet.cell(row=total_row, column=5).font = Font(bold=True)
            worksheet.cell(row=total_row, column=5).alignment = right_alignment
            worksheet.cell(row=total_row, column=5).border = border_style
            
            worksheet.cell(row=total_row, column=6, value=f"{total_importe:,.2f}")
            worksheet.cell(row=total_row, column=6).font = Font(bold=True)
            worksheet.cell(row=total_row, column=6).alignment = right_alignment
            worksheet.cell(row=total_row, column=6).border = border_style
            
            # Formatear fila de totales
            for col in range(1, 7):
                cell = worksheet.cell(row=total_row, column=col)
                if col < 4:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.border = border_style
        
        output.seek(0)
        
        # Generar nombre de archivo
        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        filename = f"Facturas_{meses[month]}_{year}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error al exportar a Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar a Excel: {str(e)}")


@router.post("/manual-create", response_model=dict)
async def create_manual_invoice(
    factura_data: ManualFacturaCreate,
    repo: FacturaRepository = Depends(get_factura_repository)
):
    """
    Crear o actualizar factura manualmente desde la sección de pendientes.
    
    Lógica:
    - Si existe factura con mismo drive_file_name en estado 'error'/'revisar': actualizar
    - Si existe factura con mismo drive_file_name en estado 'procesado': rechazar
    - Si no existe: crear nueva
    """
    try:
        import hashlib
        from datetime import datetime, date
        from decimal import Decimal
        from sqlalchemy import func
        
        drive_file_name = factura_data.drive_file_name
        logger.info(f"📝 Creando/actualizando factura manual: {drive_file_name}")
        
        with repo.db.get_session() as session:
            from src.db.models import Factura
            
            # Buscar factura existente con mismo drive_file_name
            existing_factura = session.query(Factura).filter(
                Factura.drive_file_name == drive_file_name
            ).order_by(Factura.creado_en.desc()).first()
            
            if existing_factura:
                # Si existe y está procesada, rechazar
                if existing_factura.estado == 'procesado':
                    raise HTTPException(
                        status_code=400,
                        detail=f"La factura '{drive_file_name}' ya fue procesada correctamente. No se puede crear duplicado."
                    )
                
                # Si existe y está en error/revisar, actualizar
                logger.info(f"🔄 Actualizando factura existente (ID: {existing_factura.id}, estado: {existing_factura.estado})")
                
                # Buscar o crear proveedor maestro
                from src.utils.proveedor_finder import normalizar_y_buscar_proveedor
                try:
                    resultado_proveedor = normalizar_y_buscar_proveedor(
                        nombre_raw=factura_data.proveedor_text,
                        nif=None,
                        session=session
                    )
                    existing_factura.proveedor_maestro_id = resultado_proveedor['proveedor_maestro_id']
                    logger.info(f"✅ Proveedor maestro asignado: {resultado_proveedor['nombre_canonico']}")
                except Exception as e:
                    logger.warning(f"⚠️ No se pudo asignar proveedor maestro: {e}, continuando sin él")
                
                # Calcular base_imponible automáticamente: total - impuestos
                base_imponible_calculada = Decimal(str(factura_data.importe_total)) - Decimal(str(factura_data.impuestos_total))
                
                # Calcular IVA automáticamente si base > 0
                iva_calculado = None
                if base_imponible_calculada > 0:
                    iva_calculado = (Decimal(str(factura_data.impuestos_total)) / base_imponible_calculada) * 100
                    iva_calculado = round(iva_calculado, 2)
                
                # Actualizar campos
                existing_factura.proveedor_text = factura_data.proveedor_text
                existing_factura.fecha_emision = factura_data.fecha_emision
                existing_factura.importe_total = Decimal(str(factura_data.importe_total))
                existing_factura.base_imponible = base_imponible_calculada
                existing_factura.impuestos_total = Decimal(str(factura_data.impuestos_total))
                existing_factura.iva_porcentaje = iva_calculado
                existing_factura.numero_factura = factura_data.numero_factura
                existing_factura.moneda = 'EUR'  # Siempre EUR
                existing_factura.estado = 'procesado'  # Cambiar a procesado automáticamente
                existing_factura.fecha_validacion = datetime.utcnow()  # Registrar fecha de validación
                existing_factura.error_msg = None  # Limpiar error
                existing_factura.fecha_recepcion = datetime.utcnow()
                existing_factura.actualizado_en = datetime.utcnow()
                existing_factura.extractor = 'manual'
                existing_factura.confianza = 'alta'
                
                session.commit()
                session.refresh(existing_factura)
                
                logger.info(f"✅ Factura actualizada exitosamente (ID: {existing_factura.id})")
                
                return {
                    "success": True,
                    "message": "Factura actualizada exitosamente",
                    "factura_id": existing_factura.id,
                    "action": "updated"
                }
            else:
                # No existe, crear nueva
                logger.info(f"➕ Creando nueva factura manual: {drive_file_name}")
                
                # Generar drive_file_id único para facturas manuales
                timestamp = int(datetime.utcnow().timestamp())
                hash_name = hashlib.md5(drive_file_name.encode()).hexdigest()[:8]
                drive_file_id = f"manual_{timestamp}_{hash_name}"
                
                # Buscar o crear proveedor maestro
                proveedor_maestro_id = None
                if factura_data.proveedor_text:
                    from src.utils.proveedor_finder import normalizar_y_buscar_proveedor
                    try:
                        resultado_proveedor = normalizar_y_buscar_proveedor(
                            nombre_raw=factura_data.proveedor_text,
                            nif=None,
                            session=session
                        )
                        proveedor_maestro_id = resultado_proveedor['proveedor_maestro_id']
                        logger.info(f"✅ Proveedor maestro asignado: {resultado_proveedor['nombre_canonico']}")
                    except Exception as e:
                        logger.warning(f"⚠️ No se pudo asignar proveedor maestro: {e}, continuando sin él")
                
                # Calcular base_imponible automáticamente: total - impuestos
                base_imponible_calculada = Decimal(str(factura_data.importe_total)) - Decimal(str(factura_data.impuestos_total))
                
                # Calcular IVA automáticamente si base > 0
                iva_calculado = None
                if base_imponible_calculada > 0:
                    iva_calculado = (Decimal(str(factura_data.impuestos_total)) / base_imponible_calculada) * 100
                    iva_calculado = round(iva_calculado, 2)
                
                # Crear nueva factura
                nueva_factura = Factura(
                    drive_file_id=drive_file_id,
                    drive_file_name=drive_file_name,
                    drive_folder_name='manual',
                    proveedor_maestro_id=proveedor_maestro_id,
                    proveedor_text=factura_data.proveedor_text,
                    numero_factura=factura_data.numero_factura,
                    moneda='EUR',  # Siempre EUR
                    fecha_emision=factura_data.fecha_emision,
                    fecha_recepcion=datetime.utcnow(),
                    fecha_validacion=datetime.utcnow(),  # Registrar fecha de validación
                    base_imponible=base_imponible_calculada,
                    impuestos_total=Decimal(str(factura_data.impuestos_total)),
                    iva_porcentaje=iva_calculado,
                    importe_total=Decimal(str(factura_data.importe_total)),
                    estado='procesado',  # Estado procesado automáticamente
                    extractor='manual',
                    confianza='alta',
                    revision=1,
                    creado_en=datetime.utcnow(),
                    actualizado_en=datetime.utcnow()
                )
                
                session.add(nueva_factura)
                session.commit()
                session.refresh(nueva_factura)
                
                logger.info(f"✅ Factura creada exitosamente (ID: {nueva_factura.id})")
                
                return {
                    "success": True,
                    "message": "Factura creada exitosamente",
                    "factura_id": nueva_factura.id,
                    "action": "created"
                }
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error al crear/actualizar factura manual: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar factura manual: {str(e)}"
        )
